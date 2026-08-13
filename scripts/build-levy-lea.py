# Builds src/data/levy.json: the inputs and results of Washington's Local
# Effort Assistance (LEA) formula for every district.
#
# Sources (both official OSPI, from the School Apportionment budget-prep page):
#   - 2027levyprojectiontool.xlsx - OSPI's "Enrichment Levy Pre-Ballot Approval
#     form with worksheet for estimating 2027 through 2031 Levy Authority and
#     LEA". Supplies assessed valuation, voter-approved levy amounts, LEA
#     enrollment, and the statutory assumptions that the LevyCalc sheet uses.
#     This workbook replaces the older NNNNmultiyearlacombined.xlsx naming.
#   - F-196 general fund revenue actuals - revenue code 3300 is "Local Effort
#     Assistance", i.e. the LEA dollars a district actually received.
#
# WHY CY2027 AND NOT CY2026: the CY2027 workbook's "Voter Approved" column is
# the first one that reflects the February 10 and April 28, 2026 levy
# elections - it is stamped "updated final as of june 26, 2026". The CY2026
# voter-approved amounts are byte-identical to the previous workbook, so
# staying on CY2026 would mean the site never shows what districts just voted
# on. CY2027 is also the last year OSPI models with real inputs rather than
# pure projection: assessed valuation is the actual CY2025 AV (the previous
# workbook only had a projection of it), and enrollment is March 2025-26
# AAFTE. Out-years 2028-2031 are caseload-forecast projections and are
# deliberately not published here.
#
# The formula (matches OSPI's LevyCalc sheet rows Q, R, T, V, X):
#   enrollment K       = total AAFTE - transfer out + transfer in
#   capacity per pupil = (AV * 1.50 / 1000) / (K + ALE adjustment)
#   max LEA per pupil  = threshold - capacity per pupil
#   levy rate r        = levy / AV * 1000
#   max LEA            = max LEA per pupil * (K + ALE adjustment)
#   payable LEA        = max LEA * min(r / 1.50, 1)
#
# Run: python3 scripts/build-levy-lea.py
import csv
import json
import os
import subprocess
import warnings

import openpyxl

warnings.filterwarnings('ignore')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, 'scripts', 'raw')
OUT = os.path.join(ROOT, 'src', 'data', 'levy.json')

WORKBOOK_URL = (
    'https://ospi.k12.wa.us/sites/default/files/2026-03/2027levyprojectiontool.xlsx'
)
WORKBOOK = os.path.join(RAW, '2027levyprojectiontool.xlsx')
# F-196 actuals already cached by fetch-data.mjs
REVENUES = os.path.join(RAW, 'gf-revenues-2425.csv')

CALENDAR_YEAR = 2027
# The assessed valuation OSPI uses for CY2027 is the most recent *actual*
# figure, flatlined forward (LevyCalc row J reads the "CY 2025 AV for CY 2026
# Levy" column for every projection year). So the AV column we want is not
# named for CALENDAR_YEAR.
AV_YEAR = 2026

# LevyCalc assumptions, CY2027 column C. Each is OSPI's own rounded output
# from the base values on the workbook's Sheet1 - reproduced here so the
# numbers on the site can be traced back to a formula, not a magic constant.
LEVY_CPI = 0.0623  # LevyCalc E.1, CPI for levy as of March 2026
LEA_IPD = 0.027  # LevyCalc E.2, IPD for LEA as of March 2026
LEA_THRESHOLD = 2431.77  # C. LEA Max Per Pupil = round(2221.78 * 1.027 + 150, 2)
LEA_MAX_RATE = 1.50  # D. LEA max tax rate ($ per $1,000 AV)
MAX_LEVY_PER_PUPIL = 4077.38  # A. = round(3838.26 * 1.0623, 2)
MAX_LEVY_RATE = 2.50  # B. max levy tax rate

# RCW 84.52.0531 sets a higher per-pupil levy limit for districts with 40,000+
# FTE students, and Seattle (17001) is the only one. OSPI's LevyCalc cell C18
# hardcodes exactly that split:
#   IF(A1="17001", ROUND(Sheet1!J4*(1+CPI),2), ROUND(Sheet1!J3*(1+CPI),2))
# with Sheet1!J3 = 3838.26 and Sheet1!J4 = 4505.91 (the CY2026 limits) and the
# CY2027 levy CPI of 6.23%, giving 4077.38 and 4786.63 respectively.
# The two converge at a flat $5,035 in 2031, when the statute drops the split.
LARGE_DISTRICT_CODES = ['17001']
MAX_LEVY_PER_PUPIL_LARGE = 4786.63

# The LEA inflation enhancement drops from $250 to $150 in CY2027 (2026
# session), and it is a one-year add-on rather than a new base - which is why
# the LEA threshold falls back to $2,178.78 in CY2028.
LEA_ENHANCEMENT = 150

# 'District AAFTE' sheet, 1-indexed columns (header row 6, data from row 7).
AAFTE_TOTAL = 9  # "Total SY 2025-26" - LevyCalc G.1 reads this column
AAFTE_TRANSFER_OUT = 7
AAFTE_TRANSFER_IN = 8
AAFTE_ALE_ADJ = 24  # "ALE Adj" - LEA denominator only (LevyCalc I.1)


def ensure_workbook():
    if not os.path.exists(WORKBOOK):
        os.makedirs(RAW, exist_ok=True)
        print('downloading OSPI levy/LEA workbook...')
        subprocess.run(['curl', '-sL', WORKBOOK_URL, '-o', WORKBOOK], check=True)
    return openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)


def code(value):
    return str(value).strip().zfill(5)


def num(value):
    try:
        f = float(value)
        return f if f == f else 0.0  # filter NaN
    except (TypeError, ValueError):
        return 0.0


def main():
    wb = ensure_workbook()

    # --- Assessed valuation: Data sheet, actual "CY 2025 AV for CY 2026 Levy" ---
    data_ws = wb['Data']
    rows = list(data_ws.iter_rows(min_row=2, values_only=True))
    header = rows[0]
    av_col = None
    for i, cell in enumerate(header):
        # The header carries a trailing space in this workbook, and older
        # editions suffixed "(Proj)" - match on the year phrase only.
        if cell and f'for CY {AV_YEAR} Levy' in str(cell):
            av_col = i
            break
    if av_col is None:
        raise SystemExit(f'no AV column for CY {AV_YEAR}')
    av_by_code = {}
    for row in rows[1:]:
        if row[0] and str(row[0]).strip().isdigit():
            av_by_code[code(row[0])] = num(row[av_col])

    # --- Voter-approved levy for the calendar year being modelled ---
    va_ws = wb['Voter Approved']
    va_rows = list(va_ws.iter_rows(min_row=1, values_only=True))
    va_header = va_rows[0]
    year_cols = {
        str(cell): i
        for i, cell in enumerate(va_header)
        if cell and str(cell).strip().isdigit()
    }
    latest_levy_year = str(CALENDAR_YEAR)
    if latest_levy_year not in year_cols:
        raise SystemExit(f'no voter-approved levy column for {CALENDAR_YEAR}')
    levy_col = year_cols[latest_levy_year]
    levy_by_code = {}
    for row in va_rows[1:]:
        if row[0] and str(row[0]).strip().isdigit():
            levy_by_code[code(row[0])] = num(row[levy_col])

    # --- LEA enrollment: District AAFTE, net of high/non-high transfers ---
    aafte_ws = wb['District AAFTE']
    enroll_by_code = {}
    ale_by_code = {}
    for row in aafte_ws.iter_rows(min_row=7, values_only=True):
        if not row or not row[0] or not str(row[0]).strip().isdigit():
            continue
        c = code(row[0])
        total = num(row[AAFTE_TOTAL - 1])
        transfer_out = num(row[AAFTE_TRANSFER_OUT - 1])
        transfer_in = num(row[AAFTE_TRANSFER_IN - 1])
        enroll_by_code[c] = total - transfer_out + transfer_in
        ale_by_code[c] = num(row[AAFTE_ALE_ADJ - 1])

    # --- LEA actually received: F-196 revenue code 3300 ---
    actual_lea = {}
    if os.path.exists(REVENUES):
        with open(REVENUES, newline='', encoding='utf-8-sig') as f:
            for row in csv.DictReader(f):
                if row.get('Revenue Code') == '3300' and row.get('Fund Code') == '1':
                    c = code(row['County District Code'])
                    actual_lea[c] = actual_lea.get(c, 0.0) + num(row['Amount'])
    else:
        print(f'  (no {os.path.basename(REVENUES)}; run npm run fetch-data first)')

    districts = {}
    for c, av in av_by_code.items():
        enrollment = enroll_by_code.get(c, 0.0)
        levy = levy_by_code.get(c, 0.0)
        if av <= 0 or enrollment <= 0:
            continue
        # LEA uses enrollment net of the ALE adjustment (LevyCalc K + I.1);
        # the levy cap uses K alone.
        lea_enrollment = enrollment + ale_by_code.get(c, 0.0)
        if lea_enrollment <= 0:
            continue
        # OSPI rounds at each step (LevyCalc rows Q, R, V are all ROUND(...,2)),
        # and the rounded value feeds the next row - so round here too, or the
        # totals drift by a few dollars per district against OSPI's own output.
        capacity_per_pupil = round((av * LEA_MAX_RATE / 1000) / lea_enrollment, 2)
        # Kept as a signed difference: a negative value means the district is
        # property-rich enough to clear the threshold on its own. OSPI's row R
        # floors this at zero; the site reads the signed gap, and `eligible`
        # carries the same information.
        max_lea_per_pupil = round(LEA_THRESHOLD - capacity_per_pupil, 2)
        levy_rate = (levy / av * 1000) if av else 0.0
        eligible = max_lea_per_pupil > 0
        max_lea = round(max_lea_per_pupil * lea_enrollment, 2) if eligible else 0.0
        # OSPI rounds the levy rate to 2dp before the effort ratio, then rounds
        # the ratio itself to 2dp (LevyCalc rows T and X).
        effort = (
            round(min(round(levy_rate, 2), LEA_MAX_RATE) / LEA_MAX_RATE, 2)
            if levy_rate > 0
            else 0.0
        )
        payable = max_lea * effort
        districts[c] = {
            'av': round(av),
            'enrollment': round(enrollment, 2),
            'levy': round(levy),
            'levyRate': round(levy_rate, 4),
            'capacityPerPupil': round(capacity_per_pupil, 2),
            'maxLeaPerPupil': round(max_lea_per_pupil, 2),
            'maxLea': round(max_lea),
            'payableLea': round(payable),
            'eligible': eligible,
            'actualLea': round(actual_lea.get(c, 0.0)),
        }

    out = {
        'calendarYear': CALENDAR_YEAR,
        'levyYear': int(latest_levy_year),
        'assumptions': {
            'leaThresholdPerPupil': LEA_THRESHOLD,
            'leaMaxRate': LEA_MAX_RATE,
            'maxLevyPerPupil': MAX_LEVY_PER_PUPIL,
            'maxLevyPerPupilLarge': MAX_LEVY_PER_PUPIL_LARGE,
            'largeDistrictCodes': LARGE_DISTRICT_CODES,
            'maxLevyRate': MAX_LEVY_RATE,
        },
        'sources': {
            'workbook': WORKBOOK_URL,
            'note': (
                'OSPI Enrichment Levy Pre-Ballot Approval worksheet, CY2027 '
                '(LevyCalc rows Q, R, T, V, X). Voter-approved levy amounts '
                'are final as of June 26, 2026 and include the February 10 '
                "and April 28, 2026 elections. 'actualLea' is F-196 revenue "
                'code 3300, Local Effort Assistance, school year 2024-25.'
            ),
        },
        'districts': districts,
    }

    with open(OUT, 'w') as f:
        json.dump(out, f)

    eligible_count = sum(1 for d in districts.values() if d['eligible'])
    no_levy = sum(1 for d in districts.values() if d['levy'] == 0)
    print(
        f'wrote {len(districts)} districts ({eligible_count} LEA-eligible, '
        f'{no_levy} with no approved {CALENDAR_YEAR} levy) '
        f'-> {os.path.relpath(OUT, ROOT)}'
    )
    sample = districts.get('14005')
    if sample:
        print(
            'Aberdeen check vs OSPI LevyCalc CY2027 (expect capacity 1367.92, '
            'maxLea/pupil 1063.85, rate 2.27, payable 3,056,164):'
        )
        print(
            '  ',
            {
                k: sample[k]
                for k in (
                    'capacityPerPupil',
                    'maxLeaPerPupil',
                    'levyRate',
                    'payableLea',
                    'actualLea',
                )
            },
        )


if __name__ == '__main__':
    main()
