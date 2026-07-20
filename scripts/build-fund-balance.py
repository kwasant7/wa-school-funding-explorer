# Extracts ending Total Fund Balance per district-year from fiscal.wa.gov's
# WSFCurrent.xlsm (FundBalance sheet, block 1) into scripts/raw/fund-balance.json,
# keyed by 5-digit OSPI district code. fetch-data.mjs merges this in.
import json, warnings, os
import openpyxl
warnings.filterwarnings('ignore')

RAW = os.path.join(os.path.dirname(__file__), 'raw')
wb = openpyxl.load_workbook(os.path.join(RAW, 'wsf.xlsm'), read_only=True, data_only=True)

# Title -> 5-digit code from the District sheet
code_by_title = {}
for r in wb['District'].iter_rows(min_row=2, values_only=True):
    if not r or not r[0]:
        continue
    code = str(r[0]).zfill(5)
    for t in (r[1], r[2]):
        if t:
            code_by_title[str(t).strip().lower()] = code

fb = wb['FundBalance']
rows = list(fb.iter_rows(min_row=8, values_only=True))
hdr = rows[0]
# Block 1 = first occurrence of each year (cols 1..31). Stop at the block gap.
year_col = {}
for i in range(1, len(hdr)):
    y = hdr[i]
    if y and '-' in str(y):
        ylabel = str(y)
        if ylabel not in year_col:  # keep FIRST occurrence (block 1)
            year_col[ylabel] = i
    elif year_col and (hdr[i] is None) and i > 31:
        break

WANT = ['2019-20', '2020-21', '2021-22', '2022-23', '2023-24', '2024-25']
out = {y: {} for y in WANT}
matched = 0
for r in rows[1:]:
    title = r[0]
    if not title:
        continue
    code = code_by_title.get(str(title).strip().lower())
    if not code:
        continue
    matched += 1
    for y in WANT:
        c = year_col.get(y)
        v = r[c] if c is not None else None
        if isinstance(v, (int, float)):
            out[y][code] = round(v)

with open(os.path.join(RAW, 'fund-balance.json'), 'w') as f:
    json.dump(out, f)
print(f'matched {matched} districts; 2024-25 entries: {len(out["2024-25"])}')
print('Bellevue 17405:', {y: out[y].get('17405') for y in WANT})
