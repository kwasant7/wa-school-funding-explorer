"""Builds src/data/allocation.json: how each district's STATE allocation for
2024-25 splits into the things the money actually buys.

Why this exists: the prototypical-school model is a *state* formula, so the
"how money reaches your district" walkthrough has to land on the state
allocation, not the district's total revenue. And the state allocation is not
just staff - salaries are barely half of it. The rest is benefits, materials
and operating costs, special education, transportation and the categorical
programs.

Source: OSPI's Apportionment "Final Extract" workbook for 2024-25, the same
file behind the 1191 Apportionment Summary districts receive.
  https://ospi.k12.wa.us/safs-data-files -> Final Extract - School Year 2024-25

Two sheets are used:
  App Revenue - allotments for a short, explicit list of state revenue codes
                we can positively identify as general-fund per-pupil money
                (special ed, transportation, bilingual, LAP, highly capable,
                food, LEA). This is deliberately NOT "everything under 3xxx or
                4xxx": revenue code 4100 covers both ordinary general-fund
                grants AND, for some districts, multi-year School Construction
                Assistance Program capital grants ("Capital Grant Programs"),
                which belong to the Capital Projects Fund, not the general
                fund. An earlier version of this script summed 4100 and every
                other unmatched 3xxx/4xxx code into a catch-all bucket, which
                counted that capital money as general-fund revenue and
                overstated some small districts' "state allocation" by 2-4x.
                Anything not on the explicit list is folded into the
                subtraction residual below instead of being guessed at.
  Basic Ed    - splits the single "3100 Regular Apportionment" line into
                salaries (Z375), benefits (Z384) and general-education MSOC
                (Z390). Whatever is left in 3100 after those three is reported
                as "other basic education": CTE and skills centers, Running
                Start, Alternative Learning Experience, dropout reengagement
                and substitutes, net of deductible-revenue and recovery
                adjustments that also land in 3100.

The total is anchored to the district's actual F-196 state general-fund
revenue (districts.json's rev.state, itself F-196 codes 3000+4000) rather than
summed from the Apportionment extract: "other state programs" is defined as
whatever is left after subtracting every named category above, so the total
always ties out exactly to the state revenue shown elsewhere on the site,
however the money was actually routed or whatever this extract does not itemize.

Run: python3 scripts/build-state-allocation.py
"""
import json
import os
import zipfile

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(HERE, 'raw')
DATA = os.path.join(HERE, '..', 'src', 'data')
ARCHIVE = os.path.join(RAW, '2425_extract_final.zip')
WORKBOOK = os.path.join(RAW, 'Extract Final 2024-25.xlsx')
SOURCE_URL = 'https://ospi.k12.wa.us/sites/default/files/2026-01/2425_extract-final.zip'
DISTRICTS_JSON = os.path.join(DATA, 'districts.json')

# Basic Ed column indexes (0-based) for the three pieces of 3100 we can name.
COL_CODE = 0
COL_SALARIES = 59      # Z375 TOTAL Salaries
COL_BENEFITS = 68      # Z384 TOTAL Benefits
COL_MSOC = 80          # Z390 Total GenEd MSOC

# App Revenue: an explicit allowlist of revenue codes, matched on their
# 4-digit prefix (OSPI suffixes sub-allocations, e.g. 4158 -> 415801, 415803,
# ...). Every code here is confirmed general-fund per-pupil money by its own
# label in the extract or the F-196 item dictionary. Nothing outside this list
# is summed - see the module docstring for why "everything else" is unsafe.
BUCKETS = [
    ('specialEd', ('3121', '4121', '4122', '4126')),
    ('transportation', ('4199',)),
    ('bilingual', ('4165',)),
    ('learningAssistance', ('4155',)),
    ('highlyCapable', ('4174',)),
    ('food', ('4198',)),
    ('levyEqualization', ('3300',)),
]
BASIC_ED_CODE = '3100'


def ensure_workbook():
    if os.path.exists(WORKBOOK):
        return
    if not os.path.exists(ARCHIVE):
        raise SystemExit(
            f'Missing {ARCHIVE}.\nDownload it from {SOURCE_URL}'
        )
    with zipfile.ZipFile(ARCHIVE) as zf:
        zf.extractall(RAW)


def bucket_for(code):
    """Named bucket for a state revenue code, or None if it is either not on
    the allowlist or not state general-fund money."""
    head = code[:4]
    if head == BASIC_ED_CODE:
        return 'basicEducation'
    for name, prefixes in BUCKETS:
        if head in prefixes:
            return name
    return None


def main():
    ensure_workbook()

    with open(DISTRICTS_JSON) as handle:
        rev_state_by_code = {
            d['code']: d['rev']['state']
            for d in json.load(handle)['districts']
        }

    book = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)

    # 1. Allotments by named bucket, per district.
    totals = {}
    rows = book['App Revenue'].iter_rows(values_only=True)
    next(rows)
    for row in rows:
        code = str(row[0] or '').strip().zfill(5)
        revenue = str(row[3] or '').strip()
        amount = row[7] or 0
        if not code or not revenue or not amount:
            continue
        bucket = bucket_for(revenue)
        if bucket is None:
            continue
        totals.setdefault(code, {})
        totals[code][bucket] = totals[code].get(bucket, 0) + amount

    # 2. Split the 3100 line into salaries / benefits / MSOC / residual.
    pieces = {}
    rows = book['Basic Ed'].iter_rows(values_only=True)
    next(rows)
    for row in rows:
        code = str(row[COL_CODE] or '').strip().zfill(5)
        if not code:
            continue
        pieces[code] = {
            'salaries': row[COL_SALARIES] or 0,
            'benefits': row[COL_BENEFITS] or 0,
            'msoc': row[COL_MSOC] or 0,
        }
    book.close()

    districts = {}
    skipped_no_revenue = 0
    for code, buckets in totals.items():
        rev_state = rev_state_by_code.get(code)
        if rev_state is None:
            # District not in districts.json (no F-196 filing this year, or a
            # code the two datasets don't share) - nothing to anchor the
            # total to, so it is left out rather than guessed at.
            skipped_no_revenue += 1
            continue

        basic = buckets.pop('basicEducation', 0)
        part = pieces.get(code, {'salaries': 0, 'benefits': 0, 'msoc': 0})
        salaries = part['salaries']
        benefits = part['benefits']
        msoc = part['msoc']
        # Residual keeps the pieces summing to the 3100 total exactly.
        other_basic = basic - salaries - benefits - msoc

        named_total = basic + sum(buckets.get(n, 0) for n, _ in BUCKETS)
        # Whatever the F-196 reports as state revenue beyond the named
        # categories above - the only place this script relies on a
        # subtraction rather than a positive, labeled source. See docstring.
        other_state = rev_state - named_total

        entry = {
            'salaries': round(salaries),
            'benefits': round(benefits),
            'msoc': round(msoc),
            'otherBasicEducation': round(other_basic),
            'basicEducation': round(basic),
        }
        for name, _ in BUCKETS:
            entry[name] = round(buckets.get(name, 0))
        entry['otherState'] = round(other_state)
        entry['total'] = round(rev_state)
        districts[code] = entry

    payload = {
        'schoolYear': '2024-25',
        'source': {
            'file': SOURCE_URL,
            'note': (
                "OSPI Apportionment Final Extract, 2024-25 - the data behind the "
                "1191 Apportionment Summary. Salaries (Z375), benefits (Z384) and "
                "general-education MSOC (Z390) come from the Basic Ed sheet; "
                "special education, transportation, bilingual, learning "
                "assistance, highly capable, food and levy equalization come from "
                "an explicit allowlist of App Revenue codes - not every 3xxx/4xxx "
                "code, since some (like 4100) can also carry capital-fund grants "
                "that do not belong in a general-fund total. 'Other state "
                "programs' is what is left after subtracting all of the above "
                "from the district's actual F-196 state general-fund revenue, so "
                "the total always matches that figure exactly."
            ),
        },
        'districts': districts,
    }

    out = os.path.join(DATA, 'allocation.json')
    with open(out, 'w') as handle:
        json.dump(payload, handle, separators=(',', ':'))
        handle.write('\n')

    negatives = [c for c, d in districts.items() if d['otherState'] < 0]
    big_other = [
        c for c, d in districts.items()
        if d['total'] > 0 and d['otherState'] / d['total'] > 0.25
    ]
    print(f'Wrote allocation.json for {len(districts)} districts.')
    print(f'  skipped (no matching F-196 state revenue): {skipped_no_revenue}')
    print(f'  negative "other state programs" residual: {len(negatives)}')
    if negatives:
        print(f'    {", ".join(sorted(negatives)[:12])}')
    print(f'  "other state programs" over 25% of total: {len(big_other)}')
    if big_other:
        print(f'    {", ".join(sorted(big_other)[:12])}')


if __name__ == '__main__':
    main()
